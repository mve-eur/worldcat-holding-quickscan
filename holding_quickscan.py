#!/usr/bin/env python3

import sys
import time
import base64
from pathlib import Path
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from tqdm import tqdm
from dotenv import load_dotenv
import os
import datetime

# ---------------------------
# Logging helper
# ---------------------------
def log(msg):
    print(f"[{datetime.datetime.now():%H:%M:%S}] {msg}", flush=True)

# ---------------------------
# Instellingen
# ---------------------------
TOKEN_URL    = "https://oauth.oclc.org/token"
API_BASE     = "https://americas.discovery.api.oclc.org/worldcat/search/v2"
BRIEF_BIB    = f"{API_BASE}/brief-bibs"

DELAY_SECONDS = 0.3

SCRIPT_DIR = Path(__file__).parent.resolve()
ENV_FILE = SCRIPT_DIR / ".env"

DEFAULT_OCN_COL    = "OCN"
DEFAULT_TITLE_COL  = "Titel"
DEFAULT_AUTHOR_COL = "Auteur"
DEFAULT_YEAR_COL   = "Jaar"
DEFAULT_STATUS_COL = "Status"

WORLDCAT_BASE_URL  = "https://eur.on.worldcat.org/oclc/"

# ---------------------------
# Authenticatie
# ---------------------------
class TokenManager:
    def __init__(self, wskey, secret):
        self.wskey = wskey
        self.secret = secret
        self._token = None

    def get(self):
        if not self._token:
            self._token = self._fetch()
        return self._token

    def refresh(self):
        self._token = self._fetch()
        return self._token

    def _fetch(self):
        credentials = base64.b64encode(f"{self.wskey}:{self.secret}".encode()).decode()
        resp = requests.post(
            TOKEN_URL,
            headers={
                "Authorization": f"Basic {credentials}",
                "Content-Type": "application/x-www-form-urlencoded",
            },
            data={"grant_type": "client_credentials", "scope": "wcapi"},
            timeout=30,
        )
        if resp.status_code != 200:
            log("[FOUT] Token ophalen mislukt")
            sys.exit(1)
        return resp.json()["access_token"]

# ---------------------------
# API helper
# ---------------------------
def _get(url, params, token_mgr):
    for attempt in range(2):
        try:
            resp = requests.get(
                url,
                params=params,
                headers={
                    "Authorization": f"Bearer {token_mgr.get()}",
                    "Accept": "application/json",
                },
                timeout=30,
            )
        except requests.RequestException:
            return {"status": 0, "body": {}}

        if resp.status_code == 401 and attempt == 0:
            token_mgr.refresh()
            continue

        return {"status": resp.status_code, "body": resp.json()}

    return {"status": 401, "body": {}}

# ---------------------------
# Zoekfunctie
# ---------------------------
def search_ocns(query, symbol, token_mgr):
    result = _get(
        BRIEF_BIB,
        {
            "q": query,
            "heldBySymbol": symbol,
            "limit": 50,
        },
        token_mgr,
    )

    if result.get("status") != 200:
        return []

    records = result["body"].get("briefRecords", [])
    return [str(r["oclcNumber"]) for r in records if r.get("oclcNumber")]

# ---------------------------
# Query + fallback
# ---------------------------
def process_query(title, author, year, symbol, token_mgr):

    query_parts = []

    if title:
        query_parts.append(f'ti:"{title}"')

    if author:
        query_parts.append(f'au:"{author}"')

    if year and year.isdigit():
        y = int(year)
        query_parts.append(f'yr:{y-1}-{y+1}')

    if not query_parts:
        return {"ocns": [], "status": "Geen zoekvelden"}

    query = " AND ".join(query_parts)

    ocns = search_ocns(query, symbol, token_mgr)

    if not ocns:

        if title and author:
            fallback_query = f'ti:"{title}" AND au:"{author}"'
            ocns = search_ocns(fallback_query, symbol, token_mgr)

            if ocns:
                return {"ocns": ocns, "status": f"Holding gevonden via titel/auteur ({len(ocns)} OCNs)"}

        if title:
            fallback_query = f'ti:"{title}"'
            ocns = search_ocns(fallback_query, symbol, token_mgr)

            if ocns:
                return {"ocns": ocns, "status": f"Holding gevonden via titel ({len(ocns)} OCNs)"}

        return {"ocns": [], "status": "Geen match"}

    if len(ocns) == 1:
        status = "Holding gevonden (1 OCN)"
    else:
        status = f"Holding gevonden ({len(ocns)} OCNs)"

    return {"ocns": ocns, "status": status}

# ---------------------------
# Clean helper
# ---------------------------
def clean(value):
    if pd.isna(value):
        return ""
    return str(value).strip()

# ---------------------------
# Main
# ---------------------------
def main():
    load_dotenv(dotenv_path=ENV_FILE)

    wskey  = os.getenv("WSKEY")
    secret = os.getenv("WSKEY_SECRET")
    symbol = os.getenv("INSTITUTION_SYMBOL")

    if not all([wskey, secret, symbol]):
        log("[FOUT] API-keys ontbreken")
        sys.exit(1)

    token_mgr = TokenManager(wskey, secret)

    input_path = Path(sys.argv[1])
    if len(sys.argv) > 2:
        output_path = Path(sys.argv[2])
    else:
        output_path = SCRIPT_DIR / "output.xlsx"

    if not input_path.exists():
        log(f"[FOUT] Bestand niet gevonden: {input_path}")
        sys.exit(1)

    log(f"Input bestand: {input_path}")

    df = pd.read_excel(input_path, dtype=str)

    for col in (DEFAULT_TITLE_COL, DEFAULT_AUTHOR_COL, DEFAULT_YEAR_COL):
        if col not in df.columns:
            log(f"[FOUT] Kolom '{col}' ontbreekt")
            sys.exit(1)

    df[DEFAULT_OCN_COL]    = ""
    df[DEFAULT_STATUS_COL] = ""
    df["Link"]             = ""

    total = len(df)

    multi_ocn_rows = {}
    max_ocns = 1

    log(f"{total} rijen verwerken...")

    for idx, row in enumerate(df.itertuples(index=False)):
        current = idx + 1
        total = len(df)
        percent = int((current / total) * 100)

        log(f"PROGRESS:{percent}")
        log(f"ROW:{current}/{total}")

        title  = clean(getattr(row, DEFAULT_TITLE_COL))
        author = clean(getattr(row, DEFAULT_AUTHOR_COL))
        year   = clean(getattr(row, DEFAULT_YEAR_COL))

        if not title:
            df.at[idx, DEFAULT_STATUS_COL] = "Geen titel"
            continue

        result = process_query(title, author, year, symbol, token_mgr)
        time.sleep(DELAY_SECONDS)

        ocns   = result["ocns"]
        status = result["status"]

        df.at[idx, DEFAULT_STATUS_COL] = status

        if not ocns:
            continue

        df.at[idx, DEFAULT_OCN_COL] = ", ".join(ocns)

        MAX_LINKS = 12

        if len(ocns) == 1:
            df.at[idx, "Link"] = f'=HYPERLINK("{WORLDCAT_BASE_URL}{ocns[0]}","{ocns[0]}")'
        elif len(ocns) <= MAX_LINKS:
            multi_ocn_rows[idx] = ocns
            max_ocns = max(max_ocns, len(ocns))
        else:
            df.at[idx, DEFAULT_STATUS_COL] += " (meer dan 12 OCNS. Zoek handmatig)"

    if multi_ocn_rows:
        for i in range(1, max_ocns + 1):
            df[f"Link {i}"] = ""

        for idx, ocns in multi_ocn_rows.items():
            df.at[idx, "Link"] = ""
            for i, ocn in enumerate(ocns, 1):
                df.at[idx, f"Link {i}"] = f'=HYPERLINK("{WORLDCAT_BASE_URL}{ocn}","{ocn}")'

    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    header = {cell.value: cell.column_letter for cell in ws[1]}

    for row in range(2, len(df) + 2):
        for col in header:
            if col.startswith("Link"):
                cell = ws[f"{header[col]}{row}"]
                if cell.value and "HYPERLINK" in str(cell.value):
                    cell.font = Font(color="0563C1", underline="single")

    wb.save(output_path)

    log(f"Klaar. Output: {output_path}")


if __name__ == "__main__":
    main()